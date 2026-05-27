"""Excel 解析器"""
from openpyxl import load_workbook


def parse_xlsx(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, data_only=True)
    pages = []
    page_num = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_content = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows_content.append(" | ".join(cells))
        if rows_content:
            text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows_content)
            pages.append({"page": page_num, "text": text})
            page_num += 1

    wb.close()
    return pages
